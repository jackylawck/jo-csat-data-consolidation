# processor.py
import gc
import io
import logging
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from validator import SurveyDataValidationError, validate_raw_schema


def get_column_by_keyword(df: pd.DataFrame, keyword: str) -> str:
    """動態尋找欄位，避免寫死索引造成位移報錯"""
    matched = [c for c in df.columns if keyword in str(c)]
    if not matched:
        raise SurveyDataValidationError(f"找不到包含關鍵字 '{keyword}' 的欄位！")
    return matched[0]


def generate_enterprise_report(
    raw_file_bytes, config_module, operator_id="Local_User"
) -> io.BytesIO:
    logging.info(f"開始處理問卷整合作業，操作者: {operator_id}")

    try:
        df_raw = pd.read_excel(raw_file_bytes)
        validate_raw_schema(df_raw)

        if pd.isna(df_raw.iloc[0].get("Id", None)):
            df_raw = df_raw.iloc[1:].reset_index(drop=True)

        name_col = get_column_by_keyword(df_raw, "姓名")
        comp_col = get_column_by_keyword(df_raw, "公司名稱")

        df_raw = df_raw.dropna(subset=[name_col]).reset_index(drop=True)

        # 1. 姓名標準化（若字典無則保留原名）
        df_raw["Standard_Person"] = (
            df_raw[name_col]
            .astype(str)
            .str.strip()
            .apply(
                lambda p: config_module.PERSON_NAME_MAPPING.get(p, p)
                if hasattr(config_module, "PERSON_NAME_MAPPING")
                else p
            )
        )

        # 2. 公司標準化
        def map_comp_logic(row):
            if row["Standard_Person"] == "尹錦棠":
                return "怡輝"
            raw_c = str(row[comp_col]).strip()
            return config_module.COMPANY_MAPPING.get(raw_c, raw_c)

        df_raw["Standard_Company"] = df_raw.apply(map_comp_logic, axis=1)

        # 3. 排序邏輯
        comp_rank = {
            name: i for i, name in enumerate(config_module.COMPANY_ORDER)
        }
        person_rank = {
            comp: {p: j for j, p in enumerate(plist)}
            for comp, plist in getattr(
                config_module, "PERSON_ORDER", {}
            ).items()
        }

        df_raw["comp_rank"] = (
            df_raw["Standard_Company"]
            .map(comp_rank)
            .fillna(len(config_module.COMPANY_ORDER) + 10)
        )
        df_raw["p_rank"] = df_raw.apply(
            lambda r: person_rank.get(r["Standard_Company"], {}).get(
                r["Standard_Person"], 999
            ),
            axis=1,
        )

        df_sorted = df_raw.sort_values(["comp_rank", "p_rank"]).reset_index(
            drop=True
        )

        # 4. 建立 Excel 與樣式配置
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = config_module.COMPANY_INFO["sheet_name"]
        ws.views.sheetView[0].showGridLines = True

        font_title = Font(
            name="Microsoft JhengHei", size=14, bold=True, color="1F4E79"
        )
        font_bold = Font(name="Microsoft JhengHei", size=10, bold=True)
        font_regular = Font(name="Microsoft JhengHei", size=10)
        font_red = Font(
            name="Microsoft JhengHei", size=10, color="FF0000", bold=True
        )

        align_center = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        align_left = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )

        thin_side = Side(style="thin", color="BFBFBF")
        thin_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )

        fill_header = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        fill_avg = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )
        fill_gray = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )

        def apply_style(
            cell,
            font=font_regular,
            alignment=align_center,
            fill=None,
            border=thin_border,
        ):
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if fill:
                cell.fill = fill
            if border:
                cell.border = border

        # 標題與左側題目欄
        ws.cell(1, 1, config_module.COMPANY_INFO["report_title"]).font = (
            font_title
        )
        apply_style(ws.cell(3, 1, "客戶名稱"), font=font_bold, fill=fill_gray)
        apply_style(ws.cell(5, 1, "給分人士"), font=font_bold, fill=fill_gray)

        score_cols = [
            (
                label,
                cat,
                get_column_by_keyword(df_sorted, kw),
            )
            for label, cat, kw in config_module.SCORE_QUESTIONS
        ]
        text_cols = [
            (label, get_column_by_keyword(df_sorted, kw))
            for label, kw in config_module.TEXT_QUESTIONS
        ]

        for idx, (label, cat, _) in enumerate(score_cols):
            r = 7 + idx
            apply_style(
                ws.cell(r, 1, label), font=font_regular, alignment=align_left
            )
            apply_style(
                ws.cell(r, 2, cat),
                font=font_bold,
                fill=fill_gray if cat else None,
            )

        apply_style(
            ws.cell(24, 1, "總分 (170分)"),
            font=font_bold,
            alignment=align_left,
            fill=fill_gray,
        )
        apply_style(
            ws.cell(25, 1, "各客戶代表-獨立平均分"),
            font=font_bold,
            alignment=align_left,
            fill=fill_gray,
        )
        apply_style(ws.cell(28, 1, "其他意見"), font=font_bold, fill=fill_gray)

        for idx, (label, _) in enumerate(text_cols):
            apply_style(
                ws.cell(30 + idx, 1, label),
                font=font_regular,
                alignment=align_left,
            )

        current_col = 3
        all_customer_cols = []

        # 5. 填寫每間公司與填表人資料
        for comp_name, group in df_sorted.groupby(
            "Standard_Company", sort=False
        ):
            person_cols = []
            for p_idx, (_, row_data) in enumerate(group.iterrows()):
                person_name = row_data["Standard_Person"]
                person_cols.append(current_col)
                all_customer_cols.append(current_col)

                if p_idx == 0:
                    apply_style(
                        ws.cell(3, current_col, comp_name),
                        font=font_bold,
                        fill=fill_gray,
                    )
                else:
                    apply_style(ws.cell(3, current_col), fill=fill_gray)
                apply_style(ws.cell(3, current_col + 1), fill=fill_gray)

                apply_style(
                    ws.cell(5, current_col, person_name),
                    font=font_bold,
                    fill=fill_gray,
                )
                apply_style(ws.cell(5, current_col + 1), fill=fill_gray)

                apply_style(
                    ws.cell(
                        6,
                        current_col,
                        int(config_module.COMPANY_INFO["report_year"]),
                    ),
                    font=font_bold,
                )
                apply_style(
                    ws.cell(
                        6,
                        current_col + 1,
                        int(config_module.COMPANY_INFO["last_year"]),
                    ),
                    font=font_bold,
                )

                # 填寫 1~17 題分數（少於 7 分標記紅色）
                valid_count = 0
                for q_idx, (_, _, actual_col) in enumerate(score_cols):
                    r = 7 + q_idx
                    val = row_data[actual_col]
                    try:
                        score = float(val)
                        cell_s = ws.cell(r, current_col, score)
                        if score < 7:
                            apply_style(cell_s, font=font_red)
                        else:
                            apply_style(cell_s, font=font_regular)
                        valid_count += 1
                    except:
                        cell_s = ws.cell(r, current_col, "--")
                        apply_style(cell_s, font=font_regular)

                    cell_last = ws.cell(r, current_col + 1, "--")
                    apply_style(cell_last, font=font_regular)

                # 總分 & 平均分公式
                col_let = get_column_letter(current_col)
                apply_style(
                    ws.cell(24, current_col, f"=SUM({col_let}7:{col_let}23)"),
                    font=font_bold,
                    fill=fill_gray,
                )
                apply_style(
                    ws.cell(24, current_col + 1, "--"),
                    font=font_bold,
                    fill=fill_gray,
                )

                max_pts = valid_count * 10 if valid_count > 0 else 170
                cell_p_avg = ws.cell(25, current_col, f"={col_let}24/{max_pts}")
                cell_p_avg.number_format = "0.0%"
                apply_style(cell_p_avg, font=font_bold, fill=fill_gray)
                apply_style(
                    ws.cell(25, current_col + 1, "--"),
                    font=font_bold,
                    fill=fill_gray,
                )

                # 18~23 題文字意見
                if p_idx == 0:
                    apply_style(
                        ws.cell(27, current_col, comp_name),
                        font=font_bold,
                        fill=fill_gray,
                    )
                else:
                    apply_style(ws.cell(27, current_col), fill=fill_gray)
                apply_style(ws.cell(27, current_col + 1), fill=fill_gray)

                apply_style(
                    ws.cell(28, current_col, person_name),
                    font=font_bold,
                    fill=fill_gray,
                )
                apply_style(ws.cell(28, current_col + 1), fill=fill_gray)
                apply_style(
                    ws.cell(
                        29,
                        current_col,
                        int(config_module.COMPANY_INFO["report_year"]),
                    ),
                    font=font_bold,
                )
                apply_style(
                    ws.cell(29, current_col + 1, "--"), font=font_regular
                )

                for t_idx, (_, actual_col) in enumerate(text_cols):
                    t_val = row_data[actual_col]
                    val_str = (
                        str(t_val).strip()
                        if (pd.notna(t_val) and str(t_val).strip() != "")
                        else "--"
                    )
                    apply_style(
                        ws.cell(30 + t_idx, current_col, val_str),
                        font=font_regular,
                        alignment=align_left,
                    )
                    apply_style(
                        ws.cell(30 + t_idx, current_col + 1, "--"),
                        font=font_regular,
                    )

                current_col += 2

            # 客戶每題平均分欄位
            avg_col = current_col
            apply_style(
                ws.cell(3, avg_col, f"「{comp_name}」\n就每題目的平均分"),
                font=font_bold,
                fill=fill_avg,
            )
            apply_style(ws.cell(3, avg_col + 1), fill=fill_avg)
            apply_style(ws.cell(5, avg_col), fill=fill_avg)
            apply_style(ws.cell(5, avg_col + 1), fill=fill_avg)
            apply_style(
                ws.cell(
                    6, avg_col, int(config_module.COMPANY_INFO["report_year"])
                ),
                font=font_bold,
                fill=fill_avg,
            )
            apply_style(
                ws.cell(
                    6,
                    avg_col + 1,
                    int(config_module.COMPANY_INFO["last_year"]),
                ),
                font=font_bold,
            )

            for q_idx in range(len(score_cols)):
                r = 7 + q_idx
                cls = [f"{get_column_letter(c)}{r}" for c in person_cols]
                formula = (
                    f"=AVERAGE({','.join(cls)})"
                    if len(cls) > 1
                    else f"={cls[0]}"
                )
                c_avg = ws.cell(r, avg_col, formula)
                c_avg.number_format = "0.00"
                apply_style(c_avg, font=font_bold, fill=fill_avg)
                apply_style(ws.cell(r, avg_col + 1, "--"), font=font_regular)

            apply_style(
                ws.cell(24, avg_col, "--"), font=font_bold, fill=fill_avg
            )
            apply_style(ws.cell(24, avg_col + 1, "--"), font=font_bold)
            apply_style(
                ws.cell(25, avg_col, "--"), font=font_bold, fill=fill_avg
            )
            apply_style(ws.cell(25, avg_col + 1, "--"), font=font_bold)

            current_col += 2

        # 6. 每題總平均分欄位
        tot_col = current_col
        apply_style(
            ws.cell(3, tot_col, "每題總\n平均分"),
            font=font_bold,
            fill=fill_header,
        )
        apply_style(
            ws.cell(
                6, tot_col, int(config_module.COMPANY_INFO["report_year"])
            ),
            font=font_bold,
            fill=fill_header,
        )

        for q_idx in range(len(score_cols)):
            r = 7 + q_idx
            cust_l = [f"{get_column_letter(c)}{r}" for c in all_customer_cols]
            c_tot = ws.cell(r, tot_col, f"=AVERAGE({','.join(cust_l)})")
            c_tot.number_format = "0.00"
            apply_style(c_tot, font=font_bold, fill=fill_header)

        apply_style(
            ws.cell(24, tot_col, "--"), font=font_bold, fill=fill_header
        )
        apply_style(
            ws.cell(25, tot_col, "--"), font=font_bold, fill=fill_header
        )

        # 7. 版面尺寸優化
        ws.column_dimensions["A"].width = 46
        ws.column_dimensions["B"].width = 14
        for c in range(3, current_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 13
        ws.row_dimensions[3].height = 28
        ws.row_dimensions[5].height = 24
        for r in range(7, 26):
            ws.row_dimensions[r].height = 20
        for r in range(30, 36):
            ws.row_dimensions[r].height = 42

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    finally:
        # 強制即時銷毀記憶體，確保資料隱私
        if "df_raw" in locals():
            del df_raw
        if "df_sorted" in locals():
            del df_sorted
        gc.collect()
